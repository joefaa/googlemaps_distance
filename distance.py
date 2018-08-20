from openpyxl import load_workbook
import googlemaps
import urllib
import json
import sys

def get_addresses(filename=""):
    row_number = 0
    filename=filename
    wb=load_workbook(filename=filename)
    ws= wb.active
    result_list = list()
    for row in ws.iter_rows():
        if row_number == 0:
            print(row[0].value)
            row_number += 1
        else:
            print(row[0].value)
            site = row[0].value
            address = row[2].value
            result = (site, address)
            result_list.append(result)
    return(result_list)

def get_distances(member_address, site_address ):
    origin = member_address
    destination = site_address
    key="AIzaSyA_SqXlJ7gwiWix0tjsHhWrDkxOmD_VJcg"
    units="imperial"
    params = urllib.parse.urlencode({'origins': origin, 'destinations': destination, 'key': key, 'units':units})
    url = 'https://maps.googleapis.com/maps/api/distancematrix/json?{}'.format(params)
    result = urllib.request.urlopen(url).read().decode("utf-8")
    if json.loads(result)["status"] == "OK":
        decoded = json.loads(result)["rows"][0]["elements"][0]
        print("Match OK:\n{}".format(decoded))
        distance_text = decoded['distance']['text']
        distance_value = int(decoded['distance']['value'])
        duration_text = decoded['duration']['text']
        duration_value = int(decoded['duration']['value'])
        return(distance_text,distance_value,duration_text,duration_value)
    elif json.loads(result)["status"] == "OVER_QUERY_LIMIT":
        print("Error, status: {}".format(json.loads(result)["status"]))
        sys.exit()
    else:
        print(result)
        print("Error, status: {}".format(json.loads(result)["status"]))
        return("Error","Error","Error","Error")

def matcher(input1=None,input2=None,output=None):
    input1=input1
    input2=input2
    output=output
    file = open(output, "w")
    file.write("Site, Member, Distance, Distance_Value, Duration, Duration_value\n")
    sitelist = get_addresses(input1)
    memberlist = get_addresses(input2)
    for site in sitelist:
        site_number = site[0]
        site_address = site[1]
        for member in memberlist:
            member_number = member[0]
            print("Site number: {}".format(site_number))
            print("Member number: {}".format(member_number))
            member_address = member[1]
            distance_text, travel_time, duration_text, duration_value = get_distances(member_address, site_address)
            file.write("{},{},{},{},{},{}\n".format(site_number, member_number, distance_text, duration_value, duration_text, duration_value))


matcher(input1="sites.xlsx",input2="members.xlsx",output="distancetable.csv")
